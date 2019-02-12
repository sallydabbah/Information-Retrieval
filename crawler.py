# -*- coding: utf-8 -*-

import time
import urllib.request
import re
import os.path
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import pickle


def start_crawler():
    actors_cnt = 0
    global queue_refill_cnt
    while actors_cnt < 10:
        if len(actors_queue) > 1:
            actor_id = actors_queue.pop(1)
            fetch_actor(actor_id)
            actors_cnt += 1
            save_queue()
            save_xlsx()
        else:
            cel = 'A' + str(queue_refill_cnt)
            movie_id = ws_movies[cel].value
            url = 'https://www.imdb.com/title/' + movie_id
            html = urllib.request.urlopen(url)
            page = BeautifulSoup(html, 'html.parser')
            add_movie_actors_to_queue(page)
            actors_queue[0] = queue_refill_cnt
            save_queue()
            queue_refill_cnt += 1

    print("Actors queue has {} elements".format(len(actors_queue)))


def fetch_actor(actor_id):
    if have_actor(actor_id) == 0:
        time.sleep(1)
        url = 'https://www.imdb.com/name/' + actor_id
        html = urllib.request.urlopen(url)
        page = BeautifulSoup(html, 'html.parser')

        actor_name = get_actor_name(page)

        data = get_actor_movies(page)
        data.insert(0, actor_id)
        ws_actor_movies.append(data)
        save_actor(actor_id, actor_name)
        print("Actor: {} - done".format(actor_name))
    else:
        print("Actor: {} - already exist".format(actor_id))


def get_actor_movies(page):
    movies_html = page.find_all('div', id=re.compile("^actor-"))
    movies_list = []
    for movie in movies_html:
        movie_id = movie.find('a').get('href').split('/')[2]
        movies_list.append(movie_id)
        fetch_movie(movie_id)

    movies_html = page.find_all('div', id=re.compile("^actress-"))
    for movie in movies_html:
        movie_id = movie.find('a').get('href').split('/')[2]
        movies_list.append(movie_id)
        fetch_movie(movie_id)

    return movies_list


def get_actor_name(page):
    td = page.find('td', id='overview-top')
    h1 = td.find('h1')
    name = h1.find('span', class_='itemprop').get_text()
    return name


def have_actor(actor_id):
    cols = ws_actor_movies['A']
    for cell in cols:
        if cell.value == actor_id:
            return 1
    return 0


def fetch_movie(movie_id):
    if have_movie(movie_id) == 0:
        time.sleep(1)
        url = 'https://www.imdb.com/title/' + movie_id
        html = urllib.request.urlopen(url)
        page = BeautifulSoup(html, 'html.parser')
        movie_name = get_movie_name(page).strip()
        ws_movies.append([movie_id, movie_name])
        all_movies.append(movie_id)
        print(movie_name)


def add_movie_actors_to_queue(page):
    actors_table = page.find('table', class_="cast_list")
    actors_td = actors_table.find_all('td', class_="")
    for actor in actors_td:
        actor_a = actor.find('a')
        if actor_a is not None:
            actor_id = actor_a.get('href').split('/')[2]
            actors_queue.append(actor_id)


def get_movie_name(page):
    title_div = page.find('div', class_='title_wrapper')
    movie_name = title_div.find('h1').get_text().split('<span')[0]
    return movie_name


def have_movie(movie_id):
    if movie_id in all_movies:
        return 1
    return 0


def save_xlsx():
    workbook.save('movies1.xlsx')


def save_queue():
    # write save_queue data into a file
    filename = 'actors_queue'
    outfile = open(filename, 'wb')
    pickle.dump(actors_queue, outfile)
    outfile.close()


def load_queue():
    infile = open('actors_queue', 'rb')
    global actors_queue
    actors_queue = pickle.load(infile)
    infile.close()


def save_actor(actor_id, actor_name):
    ws_actors.append([actor_id, actor_name])


def get_all_movies():
    cols = ws_movies['A']
    global all_movies
    for cell in cols:
        all_movies.append(cell.value)


# Main Code

actors_queue = [1, 'nm0908094']
queue_refill_cnt = 1
all_movies = []

if os.path.isfile('movies1.xlsx') and os.path.isfile('actors_queue'):
    workbook = load_workbook('movies1.xlsx')
    ws_actor_movies = workbook["actor_movies"]
    ws_actors = workbook["actor_name"]
    ws_movies = workbook["movie_name"]
    load_queue()
    queue_refill_cnt = actors_queue[0]
    get_all_movies()
else:
    workbook = Workbook()
    ws_actor_movies = workbook.active
    ws_actor_movies.title = "actor_movies"
    ws_actors = workbook.create_sheet("actor_name")
    ws_movies = workbook.create_sheet("movie_name")

start_crawler()

# save excel file
# save_xlsx()
